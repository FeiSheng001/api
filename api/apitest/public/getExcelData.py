#_*_ coding:utf-8 _*_
#file:getExcelData
#author:ShengFei
#email:petersat@163.com
#time:2020/6/7 16:52
"""
解析excel测试数据
1.get_Jsondata将测试数据解析成json数组格式
{sheet1:[{第一个接口数据},{第二个接口数据}],sheet2:[{第一个接口数据},{第二个接口数据}]}
2.get_listData将测试数据解析成list格式
[{第一个接口数据},{第二个接口数据},...]
"""
import openpyxl
import os

class getExcel(object):

    def __init__(self,path):
        self.path=path

    def get_jsonData(self):
        wk=openpyxl.load_workbook(self.path)
        sheetNames=wk.sheetnames
        dataDic={}
        for sheet in sheetNames:
            listData = []
            rows = wk[sheet].max_row
            cols = wk[sheet].max_column
            for row in range(2,rows+1):
                dictData = {}
                for col in range(1,cols+1):
                    #listData.append(sheet1.cell(row,col).value)
                    dictData[wk[sheet].cell(1,col).value]=wk[sheet].cell(row,col).value
                listData.append(dictData)
            dataDic.update({sheet:listData})

            for key in dataDic.keys():
                if dataDic[key]==[]:
                    dataDic.pop(key)
                    return dataDic

    def get_listData(self):
        wk = openpyxl.load_workbook(self.path)
        sheetNames = wk.sheetnames
        for sheet in sheetNames:
            rows = wk[sheet].max_row
            cols = wk[sheet].max_column
            dataList=[]
            for row in range(2,rows+1):
                dic_data={}
                for col in range(1,cols+1):
                    dic_data[wk[sheet].cell(1,col).value]=wk[sheet].cell(row,col).value
                dataList.append(dic_data)
            return dataList

if __name__ == '__main__':
    path=os.path.dirname(os.path.abspath('.'))+'/testdata/apicase.xlsx'
    a=getExcel(path).get_listData()
    print(a)

