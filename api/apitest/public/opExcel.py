#_*_ coding:utf-8 _*_
#@file:opExcel
#@author:ShengFei
#@email:petersat@163.com
#@time:2020/7/1 20:48

"""
Excel的操作：设置单元格的内容、字体样式、颜色、对齐方式
"""

import openpyxl
from openpyxl.styles import PatternFill,colors,Font,Alignment
from .logClass import Logger

#初始化一个日志器
logger=Logger().getLog()

class Excel(object):

    def __init__(self,path):
        self.path=path
        self.wk=openpyxl.load_workbook(path)

    #获取指定单元格的值
    def get_cellValue(self,sheetName,row,columm):
        cellvalue=self.wk[sheetName].cell(row,columm).value
        self.wk.close()
        return cellvalue

    #设置单元格的值
    def set_cellValue(self,sheetName,row,column,cellvalue):
        try:
            self.wk[sheetName].cell(row,column).value=cellvalue
            self.wk.save(self.path)
            self.wk.close()
        except Exception as e:
            logger.error("写入excel失败:%s"%e)

    #设置单元格背景色
    def set_cellColor(self,sheetName,row,column,color):
        self.wk[sheetName].cell(row,column).fill=PatternFill("solid",fgColor=getattr(colors,color.upper()))
        self.wk.save(self.path)
        self.wk.close()

    #设置单元格字体
    def set_cellFont(self,sheetName,row,column,fontName,fontSize,isBold):
        cellFont=Font(name=fontName,size=fontSize,bold=isBold)
        self.wk[sheetName].cell(row,column).font=cellFont
        self.wk.save(self.path)
        self.wk.close()
    #设置单元格对齐方式
    def set_cellAlign(self,sheetName,row,column,zuoyou=None,shangxia=None):
        align=Alignment(horizontal=zuoyou,vertical=shangxia)
        self.wk[sheetName].cell(row,column).alignment=align
        self.wk.save(self.path)
        self.wk.close()

if __name__ == '__main__':
    path="exceltest.xlsx"
    for name in Excel(path).wk.sheetnames:
        Excel(path).set_cellValue(name,1,2,"OK")
        Excel(path).set_cellColor(name,1,2,"RED")
        Excel(path).set_cellFont(name,1,2,"微软雅黑","15",True)
        Excel(path).set_cellAlign(name,1,2,'center','center')



